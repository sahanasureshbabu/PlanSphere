import os
import time
import pytest
from datetime import datetime
from selenium import webdriver as selenium_webdriver
from selenium.webdriver.chrome.options import Options as SeleniumOptions
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Try Appium imports
try:
    from appium import webdriver as appium_webdriver
    from appium.options.common import AppiumOptions
    APPIUM_AVAILABLE = True
except ImportError:
    APPIUM_AVAILABLE = False

# Separate result lists per framework
selenium_results = []
appium_results   = []
all_results      = []   # combined (for backward-compat reports)

# ─────────────────────────────────────────────────────────────
# DRIVER FIXTURES
# ─────────────────────────────────────────────────────────────

@pytest.fixture(scope="function")
def driver():
    """Headless Chrome WebDriver for Desktop Selenium E2E tests."""
    options = SeleniumOptions()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    drv = selenium_webdriver.Chrome(options=options)
    drv.implicitly_wait(3)
    yield drv
    drv.quit()


@pytest.fixture(scope="function")
def appium_driver():
    """Appium mobile emulation driver (Chrome mobile fallback when Appium is unavailable)."""
    loaded = False
    if APPIUM_AVAILABLE:
        try:
            opts = AppiumOptions()
            opts.set_capability("platformName", "linux")
            opts.set_capability("browserName", "Chrome")
            opts.set_capability("automationName", "Chromium")
            opts.set_capability("goog:chromeOptions", {
                "args": ["--headless=new", "--no-sandbox", "--disable-dev-shm-usage"]
            })
            drv = appium_webdriver.Remote("http://localhost:4723", options=opts)
            drv.implicitly_wait(3)
            loaded = True
            print("\nConnected to Appium Server on port 4723.")
            yield drv
            drv.quit()
        except Exception as e:
            print(f"\nAppium connection failed: {e}. Falling back to Chrome Mobile Emulation...")

    if not loaded:
        options = SeleniumOptions()
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=375,812")
        options.add_experimental_option("mobileEmulation", {
            "deviceMetrics": {"width": 375, "height": 812, "pixelRatio": 3.0},
            "userAgent": "Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.0.3 Mobile/15E148 Safari/604.1"
        })
        drv = selenium_webdriver.Chrome(options=options)
        drv.implicitly_wait(3)
        print("Initialized Chrome mobile emulation driver (Appium fallback).")
        yield drv
        drv.quit()


@pytest.fixture(autouse=True)
def test_setup_cleanup(request):
    """Automatically clears localStorage after every test to prevent state bleed."""
    yield
    for fixture_name in ["driver", "appium_driver"]:
        if fixture_name in request.fixturenames:
            try:
                drv = request.getfixturevalue(fixture_name)
                drv.execute_script("""
                    localStorage.clear();
                    if (typeof initializeDatabase === 'function') { initializeDatabase(); }
                """)
            except Exception:
                pass

# ─────────────────────────────────────────────────────────────
# RESULT COLLECTION HOOK
# ─────────────────────────────────────────────────────────────

@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    """Collects test result metadata for Excel report generation."""
    outcome = yield
    rep = outcome.get_result()

    if rep.when == "call" or (rep.when == "setup" and rep.failed):
        test_name   = item.name
        docstring   = item.obj.__doc__ or ""
        docstring   = " ".join(l.strip() for l in docstring.split("\n") if l.strip())
        description = docstring

        if hasattr(item, "callspec") and item.callspec:
            params_str = ", ".join(f"{k}={v}" for k, v in item.callspec.params.items())
            description += f" (Parameters: {params_str})" if description else f"Tested with parameters: {params_str}"

        description = description or "E2E functional verification case."
        status      = rep.outcome.upper()
        duration    = rep.duration
        error_msg   = str(rep.longreprtext) if rep.failed else ""

        # Derive module name
        parts = test_name.split("_")
        module_name = parts[1].capitalize() if len(parts) > 1 and parts[1] not in ["test", "case"] else "Core"

        # Determine framework
        is_appium = "appium" in test_name.lower() or "appium" in item.nodeid.lower()
        framework = "Appium (Mobile Web)" if is_appium else "Selenium (Desktop)"

        # Category
        nodeid_lower = item.nodeid.lower()
        if "test_security.py" in nodeid_lower:
            category = "Security"
        elif "test_vulnerability.py" in nodeid_lower:
            category = "Vulnerability"
        elif "test_load.py" in nodeid_lower:
            category = "Load"
        else:
            category = "Functional"

        record = {
            "name":        test_name,
            "framework":   framework,
            "module":      module_name,
            "description": description,
            "status":      status,
            "duration":    duration,
            "error":       error_msg,
            "category":    category,
        }
        all_results.append(record)
        if is_appium:
            appium_results.append(record)
        else:
            selenium_results.append(record)


# ─────────────────────────────────────────────────────────────
# EXCEL REPORT WRITER
# ─────────────────────────────────────────────────────────────

def _write_excel_report(report_path, title_text, results):
    if not results:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "E2E Execution Report"
    ws.views.sheetView[0].showGridLines = True

    # Colours & fonts
    navy_fill   = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    pass_fill   = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    fail_fill   = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    skip_fill   = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    title_font   = Font(name="Segoe UI", size=16, bold=True,  color="FFFFFF")
    section_font = Font(name="Segoe UI", size=11, bold=True,  color="FFFFFF")
    bold_font    = Font(name="Segoe UI", size=10, bold=True)
    regular_font = Font(name="Segoe UI", size=10)
    pass_font    = Font(name="Segoe UI", size=10, bold=True, color="065F46")
    fail_font    = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
    skip_font    = Font(name="Segoe UI", size=10, bold=True, color="374151")

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin = Border(
        left=Side(style="thin", color="E5E7EB"), right=Side(style="thin",  color="E5E7EB"),
        top=Side(style="thin",  color="E5E7EB"), bottom=Side(style="thin", color="E5E7EB"),
    )

    # Title banner
    ws.merge_cells("A1:H2")
    ws["A1"] = title_text
    ws["A1"].font      = title_font
    ws["A1"].fill      = navy_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Summary metrics
    total    = len(results)
    passed   = sum(1 for r in results if r["status"] == "PASSED")
    failed   = sum(1 for r in results if r["status"] == "FAILED")
    skipped  = sum(1 for r in results if r["status"] == "SKIPPED")
    pass_rate     = (passed / total * 100) if total > 0 else 0
    total_duration = sum(r["duration"] for r in results)

    ws["A4"] = "Run Date:"     ; ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A5"] = "Total Cases:"  ; ws["B5"] = total
    ws["C4"] = "Passed:"       ; ws["D4"] = passed
    ws["C5"] = "Failed:"       ; ws["D5"] = failed
    ws["E4"] = "Skipped:"      ; ws["F4"] = skipped
    ws["E5"] = "Pass Rate:"    ; ws["F5"] = f"{pass_rate:.1f}%"
    ws["G4"] = "Total Duration:"; ws["G5"] = f"{total_duration:.2f}s"

    for r in range(4, 6):
        for col in ["A", "C", "E", "G"]:
            ws[f"{col}{r}"].font = bold_font
        for col in ["B", "D", "F"]:
            ws[f"{col}{r}"].font = regular_font

    for r in range(4, 6):
        for ci in range(1, 9):
            ws.cell(row=r, column=ci).border = thin

    ws["F5"].font = pass_font if pass_rate == 100 else fail_font

    # Table headers
    headers = ["S.No", "Test Case Name", "Automation Driver", "Module", "Description", "Status", "Duration (s)", "Error Details"]
    header_row = 7
    ws.row_dimensions[header_row].height = 28
    for idx, hdr in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=idx)
        cell.value = hdr
        cell.font  = section_font
        cell.fill  = header_fill
        cell.alignment = center_align
        cell.border    = thin

    # Data rows
    current_row = 8
    for idx, res in enumerate(results, 1):
        ws.row_dimensions[current_row].height = 22
        ws.cell(current_row, 1, idx).alignment         = center_align
        ws.cell(current_row, 2, res["name"]).alignment = left_align
        ws.cell(current_row, 3, res["framework"]).alignment = center_align
        ws.cell(current_row, 4, res["module"]).alignment    = center_align
        ws.cell(current_row, 5, res["description"]).alignment = left_align

        sc = ws.cell(current_row, 6, res["status"])
        sc.alignment = center_align
        if res["status"] == "PASSED":
            sc.fill = pass_fill; sc.font = pass_font
        elif res["status"] == "FAILED":
            sc.fill = fail_fill; sc.font = fail_font
        else:
            sc.fill = skip_fill; sc.font = skip_font

        ws.cell(current_row, 7, round(res["duration"], 3)).alignment = center_align
        ws.cell(current_row, 8, res["error"]).alignment = left_align

        for ci in range(1, 9):
            c = ws.cell(current_row, ci)
            c.border = thin
            if ci != 6:
                c.font = regular_font
        current_row += 1

    # Column widths
    for col in ws.columns:
        max_len   = 10
        col_letter = get_column_letter(col[0].column)
        for cell in col[6:]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max_len + 3, 50)
        if col_letter in ["A", "D", "F"]:
            width = 10
        elif col_letter in ["C", "G"]:
            width = 18
        ws.column_dimensions[col_letter].width = width

    wb.save(report_path)
    print(f"Excel Report Saved successfully: {report_path}")


# ─────────────────────────────────────────────────────────────
# SESSION-END REPORT GENERATION
# ─────────────────────────────────────────────────────────────

def pytest_sessionfinish(session, exitstatus):
    """Generates Excel spreadsheets for Selenium, Appium, and combined results."""
    workspace_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

    # 1. Selenium-only report
    selenium_report_path = os.path.join(workspace_dir, "selenium_test_report.xlsx")
    sel_results = [r for r in selenium_results if r.get("category") == "Functional"]
    _write_excel_report(
        selenium_report_path,
        "PlanSphere Selenium Desktop E2E Testing Report",
        sel_results
    )

    # 2. Appium-only report
    appium_report_path = os.path.join(workspace_dir, "appium_test_report.xlsx")
    app_results = [r for r in appium_results if r.get("category") == "Functional"]
    _write_excel_report(
        appium_report_path,
        "PlanSphere Appium Mobile E2E Testing Report",
        app_results
    )

    # 3. Combined functional report
    functional_results = [r for r in all_results if r.get("category") == "Functional"]
    _write_excel_report(
        os.path.join(workspace_dir, "functional_test_report.xlsx"),
        "PlanSphere E2E Functional Testing Report (Selenium & Appium)",
        functional_results
    )

    # 4. Security report
    _write_excel_report(
        os.path.join(workspace_dir, "security_test_report.xlsx"),
        "PlanSphere Static Security & Prevention Analysis Report",
        [r for r in all_results if r.get("category") == "Security"]
    )

    # 5. Vulnerability report
    _write_excel_report(
        os.path.join(workspace_dir, "vulnerability_test_report.xlsx"),
        "PlanSphere Dependency & Secrets Vulnerability Scan Report",
        [r for r in all_results if r.get("category") == "Vulnerability"]
    )

    # 6. Load report
    _write_excel_report(
        os.path.join(workspace_dir, "load_test_report.xlsx"),
        "PlanSphere Load & Performance Stress Test Report",
        [r for r in all_results if r.get("category") == "Load"]
    )
